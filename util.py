import unittest
from openpyxl import load_workbook


class Utils(unittest.TestCase):
    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        # max number of row
        row_ct = 10
        col_ct = 2

        # 2-101 done, 1 is head title
        # 102 - 201 done
        # 202 - 301 done
        # 302 - 401 done
        # 402 - 501 done
        # 502 - 601 done
        # 602 - 701 done
        # 702 - 756 done

        for i in range(2, row_ct + 1):
            row = []
            for j in range(2, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist
